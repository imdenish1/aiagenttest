
import streamlit as st
from sentence_transformers import SentenceTransformer, util
import PyPDF2
from docx import Document
import openpyxl
import os

# Initialize the model
@st.cache_resource
def load_model():
    return SentenceTransformer('all-MiniLM-L6-v2')

model = load_model()

# File reading functions
def extract_text_from_pdf(file):
    try:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        return f"Error reading PDF file: {str(e)}"

def extract_text_from_docx(file):
    try:
        doc = Document(file)
        return "
".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        return f"Error reading Word document: {str(e)}"

def extract_text_from_excel(file):
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

def process_file(uploaded_file):
    if uploaded_file.type == "application/pdf":
        return extract_text_from_pdf(uploaded_file)
    elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return extract_text_from_docx(uploaded_file)
    elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return extract_text_from_excel(uploaded_file)
    elif uploaded_file.type.startswith("text"):
        return str(uploaded_file.read(), "utf-8")
    else:
        return "Unsupported file type."

# Upload and index files
st.title("AI Document Search Agent")

uploaded_files = st.file_uploader("Upload files for indexing", accept_multiple_files=True)
if uploaded_files:
    st.success(f"{len(uploaded_files)} file(s) uploaded successfully!")
    document_texts = []
    file_names = []
    for uploaded_file in uploaded_files:
        file_names.append(uploaded_file.name)
        document_texts.append(process_file(uploaded_file))
else:
    document_texts = []

# Semantic search
query = st.text_input("Enter your search query:")
if query and document_texts:
    embeddings = model.encode(document_texts, convert_to_tensor=True)
    query_embedding = model.encode(query, convert_to_tensor=True)
    scores = util.pytorch_cos_sim(query_embedding, embeddings)[0]
    results = sorted(zip(file_names, document_texts, scores), key=lambda x: x[2], reverse=True)

    st.subheader("Search Results:")
    for file_name, doc_text, score in results[:5]:  # Display top 5 results
        st.markdown(f"### {file_name} (Score: {score.item():.4f})")
        st.write(doc_text[:500] + "...")  # Preview the first 500 characters
        st.download_button(label=f"Download {file_name}", data=doc_text, file_name=file_name)
elif query:
    st.warning("No files indexed yet. Please upload files first.")
